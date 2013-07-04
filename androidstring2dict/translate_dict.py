#!/usr/bin/env python
# -*- coding: UTF-8 -*-

import shelve
from shelve import DbfilenameShelf
from openpyxl import Workbook, load_workbook

class TranslateDict(DbfilenameShelf):
    def __init__(self, filename):
        DbfilenameShelf.__init__(self,filename)

    def export_to_excel(self, exfilename, condition_func):
        wb = Workbook(optimized_write=True)
        ws = wb.create_sheet()

        for key in self.keys():
            value = self.__getitem__(key)
            if (condition_func(key, value)):
                ws.append((key, value))

        wb.save(exfilename)

    def _has_not_key(self, key, value):
        return not value

    def _all_items(self, key, value):
        return True

    def export_all_items(self, exfilename):
        self.export_to_excel(exfilename, self._all_items)

    def export_has_not_value(self, exfilename):
        self.export_to_excel(exfilename, self._has_not_key)

    def import_from_excel(self, exfilename, key_column, value_column):
        wb = load_workbook(filename=exfilename, use_iterators=True)
        ws = wb.get_active_sheet()

        for row in ws.iter_rows():
            try:
                member_key = row[key_column].internal_value
                member_value = row[value_column].internal_value
            except:
                pass
            else:
                if member_key is not None:
                    member_key = member_key.encode('utf-8')
                    self.__setitem__(member_key, member_value)

